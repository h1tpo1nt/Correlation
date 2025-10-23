
#с учетом временных сдвигов от -4 до +4 недель
#С УЧЕТОМ КОНКРЕТНЫХ ЗНАЧЕНИЙ ГОД И НЕДЕЛЯ

#Версия 3.1 - Автоматический поиск файлов и пропуск заголовков

import pandas as pd
import numpy as np
from scipy.stats import pearsonr
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
import glob

# ============================================================================
# НАСТРОЙКИ
# ============================================================================

# Папка с входными файлами
INPUT_FOLDER = './input'

# Папка для выходных файлов
OUTPUT_FOLDER = './output'

# Укажите имя листа, если нужно (оставьте None для первого листа)
SHEET_NAME = None  # или 'Sheet1', 'Данные' и т.д.

# Выходной файл (будет создан в OUTPUT_FOLDER)
OUTPUT_FILE = 'correlation_analysis_results.xlsx'

# Диапазон сдвигов для анализа (недели)
LAG_RANGE = range(-4, 5)  # от -4 до +4

# Колонка с названием продукта
PRODUCT_COLUMN = 'Группа продукта'

# Колонка с типом/формой продукта
FORM_COLUMN = 'mrt_archive_total[FormOfSubstance_group]'

# Колонка с локацией
LOCATION_COLUMN = 'Локация'

# Колонка с годом
YEAR_COLUMN = 'Год'

# Колонка с неделей
WEEK_COLUMN = 'Неделя'

# Метрики для анализа
METRICS = ['max', 'avg', 'min']

# ============================================================================
# ОСНОВНЫЕ ФУНКЦИИ
# ============================================================================

def find_excel_file(folder_path):
    #Находит первый Excel файл в указанной папке
    try:
        # Ищем все Excel файлы (правильный синтаксис без обратного слеша)
        excel_files = (glob.glob(os.path.join(folder_path, '*.xlsx')) + 
                      glob.glob(os.path.join(folder_path, '*.xls')))
        
        # Исключаем временные файлы Excel (начинающиеся с ~$)
        excel_files = [f for f in excel_files if not os.path.basename(f).startswith('~$')]
        
        if not excel_files:
            print(f"✗ Не найдено Excel файлов в папке '{folder_path}'")
            return None
        
        if len(excel_files) > 1:
            print(f"\\n⚠ Найдено несколько Excel файлов:")
            for idx, file in enumerate(excel_files, 1):
                print(f"  {idx}. {os.path.basename(file)}")
            print(f"\\n→ Используется первый файл: {os.path.basename(excel_files[0])}")
        
        return excel_files[0]
    except Exception as e:
        print(f"✗ Ошибка при поиске файлов: {e}")
        return None


def load_data(filename, sheet_name=None):

    #Загружает данные из Excel файла
    #ВАЖНО: Пропускает первые 2 строки, начинает читать с 3-й строки (header=2)

    try:
        # Читаем файл Excel, пропуская первые 2 строки
        # header=2 означает, что заголовки таблицы находятся в строке 3 (индексация с 0)
        excel_data = pd.read_excel(filename, sheet_name=sheet_name, header=2)
        
        # Проверяем, является ли результат словарем (несколько листов)
        if isinstance(excel_data, dict):
            print(f"\\n⚠ Обнаружено несколько листов в файле:")
            for idx, (sheet_name, sheet_df) in enumerate(excel_data.items(), 1):
                print(f"  {idx}. '{sheet_name}' - {len(sheet_df)} строк, {len(sheet_df.columns)} колонок")
            
            # Берем первый лист
            first_sheet_name = list(excel_data.keys())[0]
            df = excel_data[first_sheet_name]
            print(f"\\n→ Используется первый лист: '{first_sheet_name}'")
        else:
            df = excel_data
        
        print(f"✓ Данные успешно загружены из '{os.path.basename(filename)}'")
        print(f"  ⚠ Пропущены первые 2 строки (заголовки берутся из строки 3)")
        print(f"  Размер датасета: {len(df)} строк, {len(df.columns)} колонок")
        print(f"\\n  Доступные колонки:")
        for idx, col in enumerate(df.columns, 1):
            print(f"    {idx}. {col}")
        
        return df
    except Exception as e:
        print(f"✗ Ошибка при загрузке файла: {e}")
        import traceback
        traceback.print_exc()
        return None


def calculate_standard_correlations(pivot_data):
    #Рассчитывает стандартную матрицу корреляций без сдвигов
    # Используем только колонки с локациями
    location_cols = [col for col in pivot_data.columns if col not in ['Год', 'Неделя', 'time_index']]
    return pivot_data[location_cols].corr()


def calculate_lagged_correlations_with_time_index(df, form_type, product_name, metric='avg', 
                                                  lags=range(-4, 5),
                                                  year_col='Год',
                                                  week_col='Неделя',
                                                  location_col='Локация',
                                                  form_col='mrt_archive_total[FormOfSubstance_group]'):
    
    #Рассчитывает корреляции между локациями с временными сдвигами
    #С УЧЕТОМ КОНКРЕТНЫХ ЗНАЧЕНИЙ ГОД И НЕДЕЛЯ
    
    # Фильтруем данные по типу продукта
    df_filtered = df[df[form_col] == form_type].copy()
    
    if len(df_filtered) == 0:
        print(f"\\n⚠ ВНИМАНИЕ: Нет данных для типа '{form_type}'")
        print(f"  Доступные значения в колонке '{form_col}':")
        print(f"  {df[form_col].unique()}")
        return {}, pd.DataFrame()
    
    # Создаем временной индекс (комбинация Год и Неделя)
    df_filtered['time_index'] = df_filtered[year_col].astype(str) + '_W' + df_filtered[week_col].astype(str).str.zfill(2)
    
    # Получаем уникальные локации
    locations = sorted(df_filtered[location_col].unique())
    
    # Создаем сводную таблицу: строки = time_index, столбцы = локации
    pivot = df_filtered.pivot_table(
        values=metric,
        index=[year_col, week_col, 'time_index'],
        columns=location_col,
        aggfunc='mean'
    ).reset_index()
    
    # Сортируем по году и неделе
    pivot = pivot.sort_values([year_col, week_col]).reset_index(drop=True)
    
    print(f"\\n{'='*80}")
    print(f"Анализ корреляций для {product_name} {form_type.upper()} (метрика: {metric})")
    print(f"{'='*80}")
    print(f"Временной диапазон: {pivot[year_col].min()}-W{pivot[week_col].min():02d} до {pivot[year_col].max()}-W{pivot[week_col].max():02d}")
    print(f"Количество временных точек: {len(pivot)}")
    print(f"Локации: {', '.join(locations)}")
    
    results = {}
    
    # Для каждой пары локаций
    for i, base_loc in enumerate(locations):
        for j, compare_loc in enumerate(locations):
            if i >= j:  # Пропускаем дубликаты и самокорреляцию
                continue
            
            pair_name = f"{base_loc} vs {compare_loc}"
            results[pair_name] = {}
            
            # Для каждого временного сдвига
            for lag in lags:
                # Создаем копии данных с временными индексами
                base_data = pivot[[year_col, week_col, base_loc]].copy()
                base_data.columns = [year_col, week_col, 'value_base']
                
                compare_data = pivot[[year_col, week_col, compare_loc]].copy()
                compare_data.columns = [year_col, week_col, 'value_compare']
                
                # Применяем сдвиг к compare_data
                if lag != 0:
                    # Сдвигаем Год и Неделю для compare_data
                    compare_data['Неделя_shifted'] = compare_data[week_col] + lag
                    compare_data['Год_shifted'] = compare_data[year_col]
                    
                    # Обрабатываем переходы через границы года (52 недели в году)
                    while (compare_data['Неделя_shifted'] > 52).any():
                        mask = compare_data['Неделя_shifted'] > 52
                        compare_data.loc[mask, 'Год_shifted'] += 1
                        compare_data.loc[mask, 'Неделя_shifted'] -= 52
                    
                    while (compare_data['Неделя_shifted'] < 1).any():
                        mask = compare_data['Неделя_shifted'] < 1
                        compare_data.loc[mask, 'Год_shifted'] -= 1
                        compare_data.loc[mask, 'Неделя_shifted'] += 52
                    
                    # Объединяем: base_data по Год,Неделя = compare_data по Год_shifted,Неделя_shifted
                    merged = base_data.merge(
                        compare_data[['Год_shifted', 'Неделя_shifted', 'value_compare']],
                        left_on=[year_col, week_col],
                        right_on=['Год_shifted', 'Неделя_shifted'],
                        how='inner'
                    )
                else:
                    # Без сдвига - просто объединяем по Год и Неделя
                    merged = base_data.merge(
                        compare_data,
                        on=[year_col, week_col],
                        how='inner'
                    )
                
                # Извлекаем значения для корреляции
                base_values = merged['value_base']
                compare_values = merged['value_compare']
                
                # Удаляем NaN
                mask = ~(base_values.isna() | compare_values.isna())
                base_clean = base_values[mask]
                compare_clean = compare_values[mask]
                
                # Рассчитываем корреляцию
                if len(base_clean) >= 2:
                    corr, p_value = pearsonr(base_clean, compare_clean)
                    results[pair_name][f'lag_{lag}'] = {
                        'correlation': corr,
                        'p_value': p_value,
                        'n_points': len(base_clean)
                    }
                else:
                    results[pair_name][f'lag_{lag}'] = {
                        'correlation': np.nan,
                        'p_value': np.nan,
                        'n_points': len(base_clean)
                    }
    
    return results, pivot


def create_correlation_summary_table(corr_results):
    #Создает сводную таблицу корреляций в удобном формате
    if not corr_results:
        return pd.DataFrame()
    
    rows = []
    
    for pair, lags in corr_results.items():
        row = {'Пара локаций': pair}
        for lag_key, values in sorted(lags.items()):
            lag_num = int(lag_key.split('_')[1])
            corr = values['correlation']
            row[f'Lag_{lag_num:+d}'] = round(corr, 3) if not np.isnan(corr) else np.nan
        rows.append(row)
    
    return pd.DataFrame(rows)


def write_dataframe_with_title(writer, sheet_name, title, df, title_row=0):
    #Записывает DataFrame в Excel с заголовком в первой строке
    # Записываем DataFrame со сдвигом на 2 строки (заголовок + пустая строка)
    df.to_excel(writer, sheet_name=sheet_name, startrow=title_row + 2, index=False)
    
    # Получаем worksheet для форматирования
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Записываем заголовок в первую строку
    worksheet.cell(row=title_row + 1, column=1, value=title)
    
    # Форматируем заголовок
    title_cell = worksheet.cell(row=title_row + 1, column=1)
    title_cell.font = Font(size=12, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Объединяем ячейки заголовка на ширину таблицы
    if len(df.columns) > 1:
        worksheet.merge_cells(start_row=title_row + 1, start_column=1, 
                            end_row=title_row + 1, end_column=len(df.columns))


def create_full_excel_report(df, output_filename, 
                             product_column=PRODUCT_COLUMN,
                             form_column=FORM_COLUMN,
                             location_column=LOCATION_COLUMN,
                             year_column=YEAR_COLUMN,
                             week_column=WEEK_COLUMN,
                             metrics=METRICS,
                             lags=LAG_RANGE):
    #Создает полный Excel отчет с результатами анализа на разных листах
    
    print("\\n" + "="*80)
    print("СОЗДАНИЕ ОТЧЕТА")
    print("="*80)
    
    # Получаем уникальные типы продуктов
    unique_forms = df[form_column].unique()
    
    print(f"\\nОбнаружены типы продуктов: {', '.join([str(x) for x in unique_forms])}")
    
    writer = pd.ExcelWriter(output_filename, engine='openpyxl')
    
    sheet_num = 1
    
    # Для каждого типа продукта
    for form_type in sorted(unique_forms):
        df_form = df[df[form_column] == form_type].copy()
        
        if len(df_form) == 0:
            continue
        
        # Получаем название продукта из первой строки
        product_name = df_form[product_column].iloc[0] if product_column in df_form.columns else "Продукт"
        
        print(f"\\n{'='*80}")
        print(f"{product_name} - {form_type}: {len(df_form)} записей")
        print(f"  Локации: {', '.join(sorted(df_form[location_column].unique()))}")
        print(f"{'='*80}")
        
        # Создаем безопасное имя для листа
        safe_form_name = str(form_type).replace(' ', '_').replace('/', '_')[:15]
        safe_product_name = str(product_name).replace(' ', '_').replace('/', '_')[:10]
        
        # 1. Исходные данные
        sheet_name = f'{sheet_num}_{safe_product_name}_{safe_form_name}_Data'[:31]
        df_form.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"✓ Лист {sheet_num}: Исходные данные {product_name} {form_type}")
        sheet_num += 1
        
        # 2. Анализ корреляций
        form_results = {}
        form_pivots = {}
        
        for metric in metrics:
            results, pivot = calculate_lagged_correlations_with_time_index(
                df, 
                form_type=form_type,
                product_name=product_name,
                metric=metric,
                lags=lags,
                year_col=year_column,
                week_col=week_column,
                location_col=location_column,
                form_col=form_column
            )
            form_results[metric] = results
            form_pivots[metric] = pivot
            
            if len(pivot) > 0:
                # Сохраняем сводную таблицу по метрике
                sheet_name = f'{sheet_num}_{safe_form_name}_{metric.upper()}'[:31]
                pivot.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"✓ Лист {sheet_num}: {product_name} {form_type} {metric.upper()} по локациям")
                sheet_num += 1
        
        # 3. Корреляции без сдвига
        if len(form_pivots.get('avg', pd.DataFrame())) > 0:
            corr_0 = calculate_standard_correlations(form_pivots['avg'])
            sheet_name = f'{sheet_num}_{safe_form_name}_Corr_NoLag'[:31]
            corr_0.to_excel(writer, sheet_name=sheet_name)
            print(f"✓ Лист {sheet_num}: Корреляции {product_name} {form_type} (без сдвига)")
            sheet_num += 1
        
        # 4. Сводная таблица корреляций со сдвигами
        if form_results.get('avg'):
            lag_table = create_correlation_summary_table(form_results['avg'])
            if len(lag_table) > 0:
                sheet_name = f'{sheet_num}_{safe_form_name}_Corr_Lagged'[:31]
                lag_table.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"✓ Лист {sheet_num}: Корреляции {product_name} {form_type} со сдвигами (сводка)")
                sheet_num += 1
        
        # 5. Детальные корреляции для каждой пары
        if form_results.get('avg'):
            for pair_name, lags_data in form_results['avg'].items():
                # Создаем короткое имя листа
                safe_pair = pair_name.replace(' ', '_').replace('/', '_')[:15]
                sheet_name = f'{sheet_num}_{safe_form_name[:5]}_{safe_pair}'[:31]
                
                # Готовим данные
                detail_data = []
                for lag_key, values in sorted(lags_data.items()):
                    lag_num = int(lag_key.split('_')[1])
                    detail_data.append({
                        'Сдвиг (недели)': lag_num,
                        'Корреляция': values['correlation'],
                        'P-value': values['p_value'],
                        'Количество точек': values['n_points']
                    })
                detail_df = pd.DataFrame(detail_data)
                
                # Формируем полный заголовок с пояснением (ПРАВИЛЬНЫЙ СИНТАКСИС)
                base_loc, compare_loc = pair_name.split(' vs ')
                title = (
                    f"{product_name} {form_type.upper()}: {pair_name}\\n"
                    f"БАЗОВАЯ локация (без сдвига): {base_loc}\\n"
                    f"СРАВНИВАЕМАЯ локация (со сдвигом): {compare_loc}\\n"
                    f"Lag > 0: {compare_loc} ОПЕРЕЖАЕТ {base_loc} | Lag < 0: {compare_loc} ОТСТАЕТ от {base_loc}"
                )
                
                # Записываем с заголовком
                write_dataframe_with_title(writer, sheet_name, title, detail_df)
                
                print(f"✓ Лист {sheet_num}: Детали {product_name} {form_type} - {pair_name}")
                sheet_num += 1
    
    writer.close()
    
    print("\\n" + "="*80)
    print(f"✓ ОТЧЕТ УСПЕШНО СОЗДАН: '{output_filename}'")
    print("="*80)
    
    return output_filename


# ============================================================================
# ГЛАВНАЯ ФУНКЦИЯ
# ============================================================================

def main():
    #Основная функция для запуска анализа
    
    print("="*80)
    print("АНАЛИЗ КОРРЕЛЯЦИЙ МЕЖДУ ЛОКАЦИЯМИ ДЛЯ ПРОДУКТОВ")
    print("С УЧЕТОМ КОНКРЕТНЫХ ЗНАЧЕНИЙ ГОД И НЕДЕЛЯ (Версия 3.1)")
    print("="*80)
    
    # 1. Поиск Excel файла в папке input
    print(f"\\n→ Поиск Excel файлов в папке '{INPUT_FOLDER}'...")
    input_file = find_excel_file(INPUT_FOLDER)
    
    if input_file is None:
        print("\\n✗ Не найден Excel файл для обработки.")
        return
    
    # 2. Загрузка данных (с пропуском первых 2 строк)
    print(f"\\n→ Загрузка данных из файла...")
    df = load_data(input_file, SHEET_NAME)
    
    if df is None:
        print("\\n✗ Не удалось загрузить данные. Проверьте формат файла.")
        return
    
    # 3. Проверка наличия необходимых колонок
    required_columns = [PRODUCT_COLUMN, FORM_COLUMN, LOCATION_COLUMN, YEAR_COLUMN, WEEK_COLUMN] + METRICS
    missing_columns = [col for col in required_columns if col not in df.columns]
    
    if missing_columns:
        print(f"\\n✗ Ошибка: В данных отсутствуют колонки: {missing_columns}")
        print(f"\\n  Доступные колонки: {df.columns.tolist()}")
        return
    
    print("\\n✓ Все необходимые колонки найдены")
    
    # 4. Создание выходной папки, если её нет
    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)
        print(f"✓ Создана папка для результатов: '{OUTPUT_FOLDER}'")
    
    # 5. Формирование полного пути к выходному файлу
    output_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILE)
    
    # 6. Создание отчета
    output_file = create_full_excel_report(
        df=df,
        output_filename=output_path,
        product_column=PRODUCT_COLUMN,
        form_column=FORM_COLUMN,
        location_column=LOCATION_COLUMN,
        year_column=YEAR_COLUMN,
        week_column=WEEK_COLUMN,
        metrics=METRICS,
        lags=LAG_RANGE
    )
    
    print(f"\\n→ Откройте файл '{output_file}' в Excel для просмотра результатов")


# ============================================================================
# ЗАПУСК
# ============================================================================

if __name__ == "__main__":
    main()


# Сохраняем файл
with open('correlation_analysis.py', 'w', encoding='utf-8') as f:
    f.write(final_code)

print("="*80)
print("✅ ФИНАЛЬНЫЙ ИСПРАВЛЕННЫЙ КОД СОЗДАН")
print("="*80)
print("\nФайл: correlation_analysis.py")
print("\nВсе ошибки исправлены:")
print("✅ Убраны все обратные слеши (\\\\) для переноса строк")
print("✅ Используются скобки () для многострочных выражений")
print("✅ Исправлены f-строки")
print("✅ Правильный синтаксис для glob.glob()")
print("\nТеперь просто запустите:")
print("  python correlation_analysis.py")
