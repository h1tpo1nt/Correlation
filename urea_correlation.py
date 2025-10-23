

import pandas as pd
import numpy as np
from scipy.stats import pearsonr
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ============================================================================
# НАСТРОЙКИ
# ============================================================================

# Укажите путь к вашему Excel файлу
INPUT_FILE = './input/UREaa.xlsx'  # ИЗМЕНИТЕ НА ИМЯ ВАШЕГО ФАЙЛА

# Укажите имя листа, если нужно (оставьте None для первого листа)
SHEET_NAME = None  # или 'Sheet1', 'Данные' и т.д.

# Выходной файл
OUTPUT_FILE = 'urea_correlation_analysis_results.xlsx'

# Диапазон сдвигов для анализа (недели)
LAG_RANGE = range(-4, 5)  # от -4 до +4

# Колонка с типом продукта
FORM_COLUMN = "'mrt_archive_total'[FormOfSubstance_group]"

# Колонка с локацией
LOCATION_COLUMN = 'Локация'

# Колонка с годом
YEAR_COLUMN = 'Год'

# Колонка с неделей
WEEK_COLUMN = 'Неделя'

# Метрики для анализа
METRICS = ['max', 'avg', 'min']

# ============================================================================
# ОСНОВНЫЕ ФУНКЦИИ
# ============================================================================

def load_data(filename, sheet_name=None):
    """Загружает данные из Excel файла"""
    try:
        # Читаем файл Excel
        excel_data = pd.read_excel(filename, sheet_name=sheet_name)

        # Проверяем, является ли результат словарем (несколько листов)
        if isinstance(excel_data, dict):
            print(f"\n⚠ Обнаружено несколько листов в файле:")
            for idx, (sheet_name, sheet_df) in enumerate(excel_data.items(), 1):
                print(f"  {idx}. '{sheet_name}' - {len(sheet_df)} строк, {len(sheet_df.columns)} колонок")

            # Берем первый лист
            first_sheet_name = list(excel_data.keys())[0]
            df = excel_data[first_sheet_name]
            print(f"\n→ Используется первый лист: '{first_sheet_name}'")
        else:
            df = excel_data

        print(f"✓ Данные успешно загружены из '{filename}'")
        print(f"  Размер датасета: {len(df)} строк, {len(df.columns)} колонок")
        print(f"\n  Доступные колонки:")
        for idx, col in enumerate(df.columns, 1):
            print(f"    {idx}. {col}")

        return df
    except Exception as e:
        print(f"✗ Ошибка при загрузке файла: {e}")
        import traceback
        traceback.print_exc()
        return None


def calculate_standard_correlations(pivot_data):
    """Рассчитывает стандартную матрицу корреляций без сдвигов"""
    # Используем только колонки с локациями
    location_cols = [col for col in pivot_data.columns if col not in ['Год', 'Неделя', 'time_index']]
    return pivot_data[location_cols].corr()


def calculate_lagged_correlations_with_time_index(df, form_type, metric='avg', 
                                                  lags=range(-4, 5),
                                                  year_col='Год',
                                                  week_col='Неделя',
                                                  location_col='Локация',
                                                  form_col='mrt_archive_total[FormOfSubstance_group]'):
    """
    Рассчитывает корреляции между локациями с временными сдвигами
    С УЧЕТОМ КОНКРЕТНЫХ ЗНАЧЕНИЙ ГОД И НЕДЕЛЯ

    ЛОГИКА СДВИГОВ:
    ===============
    Для пары "Baltic vs Black Sea":
    - БАЗОВАЯ локация: Baltic (берется как есть, без сдвига)
    - СРАВНИВАЕМАЯ локация: Black Sea (к ней применяется сдвиг)

    Lag = 0:  Baltic[неделя N] vs Black Sea[неделя N]
    Lag = +1: Baltic[неделя N] vs Black Sea[неделя N+1] - Black Sea ОПЕРЕЖАЕТ Baltic на 1 неделю
    Lag = -1: Baltic[неделя N] vs Black Sea[неделя N-1] - Black Sea ОТСТАЕТ от Baltic на 1 неделю

    Положительный lag (+1, +2, +3, +4):
        → Вторая локация (Black Sea) ОПЕРЕЖАЕТ первую (Baltic)
        → Изменения во второй локации происходят РАНЬШЕ

    Отрицательный lag (-1, -2, -3, -4):
        → Вторая локация (Black Sea) ОТСТАЕТ от первой (Baltic)
        → Изменения в первой локации происходят РАНЬШЕ

    Parameters:
    - df: DataFrame с данными
    - form_type: 'granular' или 'prilled'
    - metric: метрика для анализа ('max', 'avg', 'min')
    - lags: список сдвигов для анализа
    - year_col: название колонки с годом
    - week_col: название колонки с неделей
    - location_col: название колонки с локацией
    - form_col: название колонки с типом продукта

    Returns:
    - Dictionary с результатами корреляций для каждой пары локаций
    - Pivot таблица с данными
    """

    # Фильтруем данные по типу продукта
    df_filtered = df[df[form_col] == form_type].copy()

    if len(df_filtered) == 0:
        print(f"\n⚠ ВНИМАНИЕ: Нет данных для типа '{form_type}'")
        print(f"  Доступные значения в колонке '{form_col}':")
        print(f"  {df[form_col].unique()}")
        return {}, pd.DataFrame()

    # Создаем временной индекс (комбинация Год и Неделя)
    df_filtered['time_index'] = df_filtered[year_col].astype(str) + '_W' + df_filtered[week_col].astype(str).str.zfill(2)

    # Получаем уникальные локации
    locations = sorted(df_filtered[location_col].unique())

    # Создаем сводную таблицу: строки = time_index, столбцы = локации
    pivot = df_filtered.pivot_table(
        values=metric,
        index=[year_col, week_col, 'time_index'],
        columns=location_col,
        aggfunc='mean'
    ).reset_index()

    # Сортируем по году и неделе
    pivot = pivot.sort_values([year_col, week_col]).reset_index(drop=True)

    print(f"\n{'='*80}")
    print(f"Анализ корреляций для {form_type.upper()} UREA (метрика: {metric})")
    print(f"{'='*80}")
    print(f"Временной диапазон: {pivot[year_col].min()}-W{pivot[week_col].min():02d} до {pivot[year_col].max()}-W{pivot[week_col].max():02d}")
    print(f"Количество временных точек: {len(pivot)}")
    print(f"Локации: {', '.join(locations)}")

    results = {}

    # Для каждой пары локаций
    for i, base_loc in enumerate(locations):
        for j, compare_loc in enumerate(locations):
            if i >= j:  # Пропускаем дубликаты и самокорреляцию
                continue

            pair_name = f"{base_loc} vs {compare_loc}"
            results[pair_name] = {}

            # Для каждого временного сдвига
            for lag in lags:
                # Создаем копии данных с временными индексами
                base_data = pivot[[year_col, week_col, base_loc]].copy()
                base_data.columns = [year_col, week_col, 'value_base']

                compare_data = pivot[[year_col, week_col, compare_loc]].copy()
                compare_data.columns = [year_col, week_col, 'value_compare']

                # Применяем сдвиг к compare_data
                if lag != 0:
                    # Сдвигаем Год и Неделю для compare_data
                    compare_data['Неделя_shifted'] = compare_data[week_col] + lag
                    compare_data['Год_shifted'] = compare_data[year_col]

                    # Обрабатываем переходы через границы года (52 недели в году)
                    while (compare_data['Неделя_shifted'] > 52).any():
                        mask = compare_data['Неделя_shifted'] > 52
                        compare_data.loc[mask, 'Год_shifted'] += 1
                        compare_data.loc[mask, 'Неделя_shifted'] -= 52

                    while (compare_data['Неделя_shifted'] < 1).any():
                        mask = compare_data['Неделя_shifted'] < 1
                        compare_data.loc[mask, 'Год_shifted'] -= 1
                        compare_data.loc[mask, 'Неделя_shifted'] += 52

                    # Объединяем: base_data по Год,Неделя = compare_data по Год_shifted,Неделя_shifted
                    merged = base_data.merge(
                        compare_data[['Год_shifted', 'Неделя_shifted', 'value_compare']],
                        left_on=[year_col, week_col],
                        right_on=['Год_shifted', 'Неделя_shifted'],
                        how='inner'
                    )
                else:
                    # Без сдвига - просто объединяем по Год и Неделя
                    merged = base_data.merge(
                        compare_data,
                        on=[year_col, week_col],
                        how='inner'
                    )

                # Извлекаем значения для корреляции
                base_values = merged['value_base']
                compare_values = merged['value_compare']

                # Удаляем NaN
                mask = ~(base_values.isna() | compare_values.isna())
                base_clean = base_values[mask]
                compare_clean = compare_values[mask]

                # Рассчитываем корреляцию
                if len(base_clean) >= 2:
                    corr, p_value = pearsonr(base_clean, compare_clean)
                    results[pair_name][f'lag_{lag}'] = {
                        'correlation': corr,
                        'p_value': p_value,
                        'n_points': len(base_clean)
                    }
                else:
                    results[pair_name][f'lag_{lag}'] = {
                        'correlation': np.nan,
                        'p_value': np.nan,
                        'n_points': len(base_clean)
                    }

    return results, pivot


def create_correlation_summary_table(corr_results):
    """Создает сводную таблицу корреляций в удобном формате"""
    if not corr_results:
        return pd.DataFrame()

    rows = []

    for pair, lags in corr_results.items():
        row = {'Пара локаций': pair}
        for lag_key, values in sorted(lags.items()):
            lag_num = int(lag_key.split('_')[1])
            corr = values['correlation']
            row[f'Lag_{lag_num:+d}'] = round(corr, 3) if not np.isnan(corr) else np.nan
        rows.append(row)

    return pd.DataFrame(rows)


def write_dataframe_with_title(writer, sheet_name, title, df, title_row=0):
    """
    Записывает DataFrame в Excel с заголовком в первой строке

    Parameters:
    - writer: ExcelWriter объект
    - sheet_name: имя листа
    - title: текст заголовка
    - df: DataFrame для записи
    - title_row: номер строки для заголовка (0-indexed)
    """
    # Записываем DataFrame со сдвигом на 2 строки (заголовок + пустая строка)
    df.to_excel(writer, sheet_name=sheet_name, startrow=title_row + 2, index=False)

    # Получаем worksheet для форматирования
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    # Записываем заголовок в первую строку
    worksheet.cell(row=title_row + 1, column=1, value=title)

    # Форматируем заголовок
    title_cell = worksheet.cell(row=title_row + 1, column=1)
    title_cell.font = Font(size=12, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Объединяем ячейки заголовка на ширину таблицы
    if len(df.columns) > 1:
        worksheet.merge_cells(start_row=title_row + 1, start_column=1, 
                            end_row=title_row + 1, end_column=len(df.columns))


def create_full_excel_report(df, output_filename, 
                             form_column=FORM_COLUMN,
                             location_column=LOCATION_COLUMN,
                             year_column=YEAR_COLUMN,
                             week_column=WEEK_COLUMN,
                             metrics=METRICS,
                             lags=LAG_RANGE):
    """
    Создает полный Excel отчет с результатами анализа на разных листах
    """

    print("\n" + "="*80)
    print("СОЗДАНИЕ ОТЧЕТА")
    print("="*80)

    writer = pd.ExcelWriter(output_filename, engine='openpyxl')

    # Разделяем данные
    df_granular = df[df[form_column] == 'granular'].copy()
    df_prilled = df[df[form_column] == 'prilled'].copy()

    print(f"\nGranular Urea: {len(df_granular)} записей")
    if len(df_granular) > 0:
        print(f"  Локации: {', '.join(sorted(df_granular[location_column].unique()))}")

    print(f"\nPrilled Urea: {len(df_prilled)} записей")
    if len(df_prilled) > 0:
        print(f"  Локации: {', '.join(sorted(df_prilled[location_column].unique()))}")

    # 1. Исходные данные - Granular
    if len(df_granular) > 0:
        df_granular.to_excel(writer, sheet_name='1_Granular_Raw_Data', index=False)
        print("\n✓ Лист 1: Исходные данные Granular")

    # 2. Исходные данные - Prilled
    if len(df_prilled) > 0:
        df_prilled.to_excel(writer, sheet_name='2_Prilled_Raw_Data', index=False)
        print("✓ Лист 2: Исходные данные Prilled")

    # 3-4. Анализ для Granular
    if len(df_granular) > 0:
        print("\n→ Расчет корреляций для Granular...")
        granular_results = {}
        granular_pivots = {}

        for metric in metrics:
            results, pivot = calculate_lagged_correlations_with_time_index(
                df, 
                form_type='granular',
                metric=metric,
                lags=lags,
                year_col=year_column,
                week_col=week_column,
                location_col=location_column,
                form_col=form_column
            )
            granular_results[metric] = results
            granular_pivots[metric] = pivot

            if len(pivot) > 0:
                # Сохраняем сводную таблицу по метрике
                pivot.to_excel(writer, sheet_name=f'3_Granular_{metric.upper()}', index=False)
                print(f"✓ Лист 3: Granular {metric.upper()} по локациям")

    # 5-6. Анализ для Prilled
    if len(df_prilled) > 0:
        print("\n→ Расчет корреляций для Prilled...")
        prilled_results = {}
        prilled_pivots = {}

        for metric in metrics:
            results, pivot = calculate_lagged_correlations_with_time_index(
                df, 
                form_type='prilled',
                metric=metric,
                lags=lags,
                year_col=year_column,
                week_col=week_column,
                location_col=location_column,
                form_col=form_column
            )
            prilled_results[metric] = results
            prilled_pivots[metric] = pivot

            if len(pivot) > 0:
                # Сохраняем сводную таблицу по метрике
                pivot.to_excel(writer, sheet_name=f'4_Prilled_{metric.upper()}', index=False)
                print(f"✓ Лист 4: Prilled {metric.upper()} по локациям")

    # 7. Корреляции без сдвига - Granular
    if len(df_granular) > 0 and len(granular_pivots.get('avg', pd.DataFrame())) > 0:
        granular_corr_0 = calculate_standard_correlations(granular_pivots['avg'])
        granular_corr_0.to_excel(writer, sheet_name='5_Granular_Corr_NoLag')
        print("✓ Лист 5: Корреляции Granular (без сдвига)")

    # 8. Корреляции без сдвига - Prilled
    if len(df_prilled) > 0 and len(prilled_pivots.get('avg', pd.DataFrame())) > 0:
        prilled_corr_0 = calculate_standard_correlations(prilled_pivots['avg'])
        prilled_corr_0.to_excel(writer, sheet_name='6_Prilled_Corr_NoLag')
        print("✓ Лист 6: Корреляции Prilled (без сдвига)")

    # 9. Сводная таблица корреляций со сдвигами - Granular
    if len(df_granular) > 0 and granular_results.get('avg'):
        granular_lag_table = create_correlation_summary_table(granular_results['avg'])
        if len(granular_lag_table) > 0:
            granular_lag_table.to_excel(writer, sheet_name='7_Granular_Corr_Lagged', index=False)
            print("✓ Лист 7: Корреляции Granular со сдвигами (сводка)")

    # 10. Сводная таблица корреляций со сдвигами - Prilled
    if len(df_prilled) > 0 and prilled_results.get('avg'):
        prilled_lag_table = create_correlation_summary_table(prilled_results['avg'])
        if len(prilled_lag_table) > 0:
            prilled_lag_table.to_excel(writer, sheet_name='8_Prilled_Corr_Lagged', index=False)
            print("✓ Лист 8: Корреляции Prilled со сдвигами (сводка)")

    # 11+. Детальные корреляции для каждой пары - Granular
    sheet_num = 9
    if len(df_granular) > 0 and granular_results.get('avg'):
        for pair_name, lags_data in granular_results['avg'].items():
            # Создаем короткое имя листа
            safe_name = pair_name.replace(' ', '_').replace('/', '_')[:20]
            sheet_name = f'{sheet_num}_G_{safe_name}'[:31]

            # Готовим данные
            detail_data = []
            for lag_key, values in sorted(lags_data.items()):
                lag_num = int(lag_key.split('_')[1])
                detail_data.append({
                    'Сдвиг (недели)': lag_num,
                    'Корреляция': values['correlation'],
                    'P-value': values['p_value'],
                    'Количество точек': values['n_points']
                })
            detail_df = pd.DataFrame(detail_data)

            # Формируем полный заголовок с пояснением
            base_loc, compare_loc = pair_name.split(' vs ')
            title = f"GRANULAR UREA: {pair_name}\n" \
                   f"БАЗОВАЯ локация (без сдвига): {base_loc}\n" \
                   f"СРАВНИВАЕМАЯ локация (со сдвигом): {compare_loc}\n" \
                   f"Lag > 0: {compare_loc} ОПЕРЕЖАЕТ {base_loc} | Lag < 0: {compare_loc} ОТСТАЕТ от {base_loc}"

            # Записываем с заголовком
            write_dataframe_with_title(writer, sheet_name, title, detail_df)

            print(f"✓ Лист {sheet_num}: Детали Granular - {pair_name}")
            sheet_num += 1

    # Детальные корреляции для каждой пары - Prilled
    if len(df_prilled) > 0 and prilled_results.get('avg'):
        for pair_name, lags_data in prilled_results['avg'].items():
            # Создаем короткое имя листа
            safe_name = pair_name.replace(' ', '_').replace('/', '_')[:20]
            sheet_name = f'{sheet_num}_P_{safe_name}'[:31]

            # Готовим данные
            detail_data = []
            for lag_key, values in sorted(lags_data.items()):
                lag_num = int(lag_key.split('_')[1])
                detail_data.append({
                    'Сдвиг (недели)': lag_num,
                    'Корреляция': values['correlation'],
                    'P-value': values['p_value'],
                    'Количество точек': values['n_points']
                })
            detail_df = pd.DataFrame(detail_data)

            # Формируем полный заголовок с пояснением
            base_loc, compare_loc = pair_name.split(' vs ')
            title = f"PRILLED UREA: {pair_name}\n" \
                   f"БАЗОВАЯ локация (без сдвига): {base_loc}\n" \
                   f"СРАВНИВАЕМАЯ локация (со сдвигом): {compare_loc}\n" \
                   f"Lag > 0: {compare_loc} ОПЕРЕЖАЕТ {base_loc} | Lag < 0: {compare_loc} ОТСТАЕТ от {base_loc}"

            # Записываем с заголовком
            write_dataframe_with_title(writer, sheet_name, title, detail_df)

            print(f"✓ Лист {sheet_num}: Детали Prilled - {pair_name}")
            sheet_num += 1

    writer.close()

    print("\n" + "="*80)
    print(f"✓ ОТЧЕТ УСПЕШНО СОЗДАН: '{output_filename}'")
    print("="*80)

    return output_filename


# ============================================================================
# ГЛАВНАЯ ФУНКЦИЯ
# ============================================================================

def main():
    """Основная функция для запуска анализа"""

    print("="*80)
    print("АНАЛИЗ КОРРЕЛЯЦИЙ МЕЖДУ ЛОКАЦИЯМИ ДЛЯ GRANULAR И PRILLED UREA")
    print("С УЧЕТОМ КОНКРЕТНЫХ ЗНАЧЕНИЙ ГОД И НЕДЕЛЯ (Версия 2.2)")
    print("="*80)

    # 1. Загрузка данных
    print("\n→ Загрузка данных...")
    df = load_data(INPUT_FILE, SHEET_NAME)

    if df is None:
        print("\n✗ Не удалось загрузить данные. Проверьте путь к файлу.")
        return

    # 2. Проверка наличия необходимых колонок
    required_columns = [FORM_COLUMN, LOCATION_COLUMN, YEAR_COLUMN, WEEK_COLUMN] + METRICS
    missing_columns = [col for col in required_columns if col not in df.columns]

    if missing_columns:
        print(f"\n✗ Ошибка: В данных отсутствуют колонки: {missing_columns}")
        print(f"\n  Доступные колонки: {df.columns.tolist()}")
        return

    print("\n✓ Все необходимые колонки найдены")

    # 3. Создание отчета
    output_file = create_full_excel_report(
        df=df,
        output_filename=OUTPUT_FILE,
        form_column=FORM_COLUMN,
        location_column=LOCATION_COLUMN,
        year_column=YEAR_COLUMN,
        week_column=WEEK_COLUMN,
        metrics=METRICS,
        lags=LAG_RANGE
    )

    print(f"\n→ Откройте файл '{output_file}' в Excel для просмотра результатов")


# ============================================================================
# ЗАПУСК
# ============================================================================

if __name__ == "__main__":
    main()
